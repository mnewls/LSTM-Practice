from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains

# open each - grab date, job title, location - these will be different cols in the excel.
# 

from openpyxl import Workbook

import html5lib

from selenium.webdriver.support.ui import Select

# to find links
from bs4 import BeautifulSoup
import urllib.request

import time # to sleep

wb = Workbook()
ws = wb.active
ws.title = "Jobs"

ws['B1'] = "Day"
ws['C1'] = "Month"
ws['D1'] = "Year"
ws['E1'] = "WB1"
ws['F1'] = "WB2"
ws['G1'] = "WB3"
ws['H1'] = "WB4"
ws['I1'] = "WB5"
ws['J1'] = "PB"
ws['K1'] = "prize"

def get_page_info(driver):
    count = 2

    for url_year in range(1992, 2021, 1):

        url_str = r'https://www.lotto.net/powerball/numbers/' + str(url_year)

        #print(url_str)

        driver.get(url_str)

        time.sleep(2)

        #print(driver.find_element_by_xpath("//*[@id='content']/div[1]/div[1]/div[1]/text()").getText())

        page_source = driver.page_source
        soup = BeautifulSoup(page_source, "html5lib")

        list_dates = []

        is_odd = 1
        list_days = []
        list_year = []
        list_month = []

        for date in soup.findAll("div", {"class": "date"}):
            date_str = date.text[14:len(date.text) - 5]
            
            #if is_odd % 2 != 0:
                #date_str = date_str[1:len(date_str)]

            month = date_str[0:4]

            #print(month)

            if "Jan" in month:
                month_num = 1
            elif "Feb" in month:
                month_num = 2
            elif "Mar" in month:
                month_num = 3
            elif "Apr" in month:
                month_num = 4
            elif 'May' in month:
                month_num = 5
            elif 'Jun' in month:
                month_num = 6
            elif 'Jul' in month:
                month_num = 7
            elif 'Aug' in month:
                month_num = 8
            elif 'Sep' in month:
                month_num = 9
            elif 'Oct' in month:
                month_num = 10
            elif 'Nov' in month:
                month_num = 11
            else:
                month_num = 12

            print(month_num)

            list_month.append(month_num)

            #print(month_num)

            #print(date_str)

            date_nums = []

            date_cleaned = date_str.replace('t', '')
            date_cleaned = date_cleaned.replace('h', '')
            date_cleaned = date_cleaned.replace('r', '')
            date_cleaned = date_cleaned.replace('s', '')
            date_cleaned = date_cleaned.replace('n', '')
            date_cleaned = date_cleaned.replace('d', '')

            for word in date_cleaned.split():
                if word.isdigit():
                    date_nums.append(int(word))

            list_days.append(date_nums[0])

            list_year.append(date_nums[1])

            is_odd+=1


            #date_str = date_str.replace('t', '')
            #date_str = date_str.replace('h', '')

            list_dates.append(date_str)

        list_len = len(list_dates)

        #print(len(list_month))

        #print(list_days)
        #print(len(list_days))

        #print(list_year)
        #print(len(list_year))

        #list_jackpots = soup.findAll("div", {"class": "jackpot"})

        list_jackpots = []

        for jackpot in soup.findAll("div", {"class": "jackpot"}):
            jackpot_str = jackpot.text[31:len(jackpot.text) - 21]

            jackpot_str = jackpot_str.replace('t', '')
            jackpot_str = jackpot_str.replace('n', '')
            jackpot_str = jackpot_str.replace('\'', '')

            list_jackpots.append(jackpot_str)

        #print(list_jackpots)
        #print(list_jackpots)
        #list_nums = soup.findAll("li", {"class": "ball ball"})

        ball_num_list = []

        for ball in soup.findAll("li", {"class": "ball ball"}):
            ball_num = ball.text

            ball_num_list.append(ball_num)
        
        pwr_ball_num_list = []

        for pwr_ball in soup.findAll("li", {"class": "ball powerball"}):
            pwr_ball_num = pwr_ball.text[0:len(pwr_ball.text)-9]

            pwr_ball_num_list.append(pwr_ball_num)

        #print(len(pwr_ball_num_list))

        WB1_list = ball_num_list[0:len(ball_num_list):5]
        WB2_list = ball_num_list[1:len(ball_num_list):5]
        WB3_list = ball_num_list[2:len(ball_num_list):5]
        WB4_list = ball_num_list[3:len(ball_num_list):5]
        WB5_list = ball_num_list[4:len(ball_num_list):5]

        #print(PB_list)
        #print(list_len)

        #print(WB1_list)

        #print(ball_num_list)

        #for i in num_draws

        for i in range(list_len):
            day_place = 'B' + str(count)
            month_place = 'C' + str(count)
            year_place = 'D' + str(count)
            WB1_place = 'E' + str(count)
            WB2_place = 'F' + str(count)
            WB3_place = 'G' + str(count)
            WB4_place = 'H' + str(count)
            WB5_place = 'I' + str(count)
            PB_place = 'J' + str(count)
            jackpot_place = 'K' + str(count)

            ws[day_place] = list_days[i]
            ws[month_place] = list_month[i]
            ws[year_place] = list_year[i]
            ws[WB1_place] = WB1_list[i]
            ws[WB2_place] = WB2_list[i]
            ws[WB3_place] = WB3_list[i]
            ws[WB4_place] = WB4_list[i]
            ws[WB5_place] = WB5_list[i]
            ws[PB_place] = pwr_ball_num_list[i]
            ws[jackpot_place] = list_jackpots[i]

            count += 1




def get_info():   

    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")

    driver = webdriver.Chrome(executable_path=r'C:\Users\Michael\Desktop\Automate Application\chromedriver.exe', chrome_options=options)

    get_page_info(driver)
        
    wb.save('test_workbook.xlsx')


get_info()